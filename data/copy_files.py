import csv
import os
import pathlib
import shutil

from PIL.Image import Resampling
from openpyxl.styles import PatternFill
from voc_tools.constants import VOC_IMAGES
from voc_tools.reader import list_dir

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage


def makedirs(path):
    try:
        os.makedirs(path)
    except:
        ...


def copy():
    dataset = pathlib.Path(r"C:\Users\dndlssardar\OneDrive - Smiths Group\Documents\Projects\Dataset\Sixray_easy")
    destination = pathlib.Path(r"./sixray_500_GK")
    
    makedirs(str(destination / "train" / "images"))
    makedirs(str(destination / "test" / "images"))
    makedirs(str(destination / "train" / "texts"))
    makedirs(str(destination / "test" / "texts"))
    
    train_filenames = list(list_dir(str(dataset / "train"), dir_flag=VOC_IMAGES, fullpath=False))
    test_filenames = list(list_dir(str(dataset / "test"), dir_flag=VOC_IMAGES, fullpath=False))
    
    with open("data_images_all.csv", "r") as fp:
        for line in fp:
            filename = line.split(',')[0].strip()
            if filename in train_filenames:
                shutil.copyfile(dataset / "train" / "JPEGImages" / filename,
                                destination / "train" / "images" / filename)
                print(".", end="")
            if filename in test_filenames:
                shutil.copyfile(dataset / "test" / "JPEGImages" / filename,
                                destination / "train" / "images" / filename)
                
                print(".", end="")
        print("Copy done.")


def generate_excel_with_thumbnails(image_paths, save_path, file_names_to_highlight=()):
    # Create a new workbook and get the active sheet
    workbook = Workbook()
    sheet = workbook.active
    os.makedirs("./temp_files", exist_ok=True)
    
    # Set the column widths to accommodate the thumbnails and file names
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 40
    
    # Set the headers for the columns
    sheet['A1'] = 'Thumbnail'
    sheet['B1'] = 'File Name'
    sheet['C1'] = 'Is in 500 collection?'
    sheet['D1'] = 'Map File Name'
    
    # Iterate through the image paths and insert thumbnails and file names into the sheet
    for i, image_path in enumerate(image_paths):
        # Open the image using PIL
        img = PILImage.open(image_path)
        
        # Calculate the aspect ratio to maintain the original proportions
        aspect_ratio = img.width / img.height
        
        # Calculate the corresponding height to maintain the aspect ratio
        thumbnail_height = 100  # int(thumbnail_width / aspect_ratio)
        
        # Set the desired width for the thumbnail (you can adjust this value)
        thumbnail_width = int(thumbnail_height * aspect_ratio)
        
        # Resize the image to the desired thumbnail size
        img = img.resize((thumbnail_width, thumbnail_height), Resampling.LANCZOS)
        
        # Convert the PIL image to an openpyxl image
        temp_path = "./temp_files/temp_" + os.path.basename(image_path)
        img.save(temp_path)
        img = Image(temp_path)
        
        # Add the thumbnail image and file name to the sheet
        filename = os.path.basename(image_path)
        sheet.add_image(img, f'A{i + 2}')
        sheet[f'B{i + 2}'] = filename  # Extract the file name from the image path
        if filename in file_names_to_highlight:
            highlight_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            sheet[f'A{i + 2}'].fill = highlight_fill
            sheet[f'B{i + 2}'].fill = highlight_fill
            sheet[f'C{i + 2}'].fill = highlight_fill
            sheet[f'C{i + 2}'] = 1
        else:
            sheet[f'C{i + 2}'] = 0
        print(".", end="")
    print("Saving into file...", end="")
    # Save the workbook as an Excel file
    workbook.save(save_path)
    print("Cleaning...", end="")
    shutil.rmtree("./temp_files", ignore_errors=True, onerror=lambda: print("Cleaning failed", end=""))
    print("File Saved")


def generate_excel_form_csv_with_thumbnails(image_paths, save_path, csv_path):
    # Create a new workbook and get the active sheet
    workbook = Workbook()
    sheet = workbook.active
    os.makedirs("./temp_files", exist_ok=True)
    
    with open(csv_path, "r") as fp:
        csv.reader(fp)
    
    # Set the column widths to accommodate the thumbnails and file names
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 40
    
    # Set the headers for the columns
    sheet['A1'] = 'Thumbnail'
    sheet['B1'] = 'File Name'
    sheet['C1'] = 'Is in 500 collection?'
    sheet['D1'] = 'Map File Name'
    
    # Iterate through the image paths and insert thumbnails and file names into the sheet
    for i, image_path in enumerate(image_paths):
        # Open the image using PIL
        img = PILImage.open(image_path)
        
        # Calculate the aspect ratio to maintain the original proportions
        aspect_ratio = img.width / img.height
        
        # Calculate the corresponding height to maintain the aspect ratio
        thumbnail_height = 100  # int(thumbnail_width / aspect_ratio)
        
        # Set the desired width for the thumbnail (you can adjust this value)
        thumbnail_width = int(thumbnail_height * aspect_ratio)
        
        # Resize the image to the desired thumbnail size
        img = img.resize((thumbnail_width, thumbnail_height), Resampling.LANCZOS)
        
        # Convert the PIL image to an openpyxl image
        temp_path = "./temp_files/temp_" + os.path.basename(image_path)
        img.save(temp_path)
        img = Image(temp_path)
        
        # Add the thumbnail image and file name to the sheet
        filename = os.path.basename(image_path)
        sheet.add_image(img, f'A{i + 2}')
        sheet[f'B{i + 2}'] = filename  # Extract the file name from the image path
        if filename in file_names_to_highlight:
            highlight_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            sheet[f'A{i + 2}'].fill = highlight_fill
            sheet[f'B{i + 2}'].fill = highlight_fill
            sheet[f'C{i + 2}'].fill = highlight_fill
            sheet[f'C{i + 2}'] = 1
        else:
            sheet[f'C{i + 2}'] = 0
        print(".", end="")
    print("Saving into file...", end="")
    # Save the workbook as an Excel file
    workbook.save(save_path)
    print("Cleaning...", end="")
    shutil.rmtree("./temp_files", ignore_errors=True, onerror=lambda: print("Cleaning failed", end=""))
    print("File Saved")


def generate_excel_with_thumbnails_pil(image_paths, save_path, mark_file_list, caption_dict, original_index_dict):
    # Create a new workbook and get the active sheet
    workbook = Workbook()
    sheet = workbook.active
    os.makedirs("./temp_files", exist_ok=True)
    
    # Set the column widths to accommodate the thumbnails and file names
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 40
    
    # Set the headers for the columns
    sheet['A1'] = 'Thumbnail'
    sheet['B1'] = 'File Name'
    sheet['C1'] = 'Index'
    sheet['D1'] = 'Captioned?'
    sheet['E1'] = 'Caption-1'
    sheet['F1'] = 'Is occluded-1?'
    sheet['G1'] = 'Is error-1?'
    sheet['H1'] = 'Original Index'
    
    # Iterate through the image paths and insert thumbnails and file names into the sheet
    for i, image_path in enumerate(image_paths):
        # Open the image using PIL
        img = PILImage.open(image_path)
        
        # Calculate the aspect ratio to maintain the original proportions
        aspect_ratio = img.width / img.height
        
        # Set the desired width for the thumbnail (you can adjust this value)
        thumbnail_width = 100
        
        # Calculate the corresponding height to maintain the aspect ratio
        thumbnail_height = int(thumbnail_width / aspect_ratio)
        
        # Resize the image to the desired thumbnail size
        img = img.resize((thumbnail_width, thumbnail_height), Resampling.LANCZOS)
        
        # Convert the PIL image to an openpyxl image
        temp_path = "./temp_files/temp_" + os.path.basename(image_path)
        img.save(temp_path)
        img = Image(temp_path)
        
        # Add the thumbnail image and file name to the sheet
        filename = os.path.basename(image_path)
        sheet.add_image(img, f'A{i + 2}')
        sheet[f'B{i + 2}'] = filename  # Extract the file name from the image path
        sheet[f'C{i + 2}'] = i + 1  # File
        sheet[f'D{i + 2}'] = 1 if filename in mark_file_list else 0
        sheet[f'E{i + 2}'] = caption_dict.get(filename, "")
        sheet[f'F{i + 2}'] = 0
        sheet[f'G{i + 2}'] = 0
        sheet[f'H{i + 2}'] = original_index_dict.get(filename, 0)
        print(".", end="")
    print("Saving into file...", end="")
    # Save the workbook as an Excel file
    workbook.save(save_path)
    print("Cleaning...", end="")
    shutil.rmtree("./temp_files", ignore_errors=True, onerror=lambda: print("Cleaning failed", end=""))
    print("File Saved")


def main():
    dataset = pathlib.Path(r"C:\Users\dndlssardar\OneDrive - Smiths Group\Documents\Projects\Dataset\Sixray_easy")
    destination = pathlib.Path(r"archive/dataset_500_to_2300_map2.xlsx")
    
    train_filenames = list(list_dir(str(dataset / "train"), dir_flag=VOC_IMAGES, fullpath=False))
    test_filenames = list(list_dir(str(dataset / "test"), dir_flag=VOC_IMAGES, fullpath=False))
    image_paths = []
    # with open("caption_soumen_clean.csv", "r") as fp:
    #     dicta = {f: c for f, c in map(lambda x: (x[:11].strip(","), x[11:].strip().strip(",").strip("\"")), fp)}
    #
    # with open("dataset_easy_lookup-backup.csv", "r") as fp:
    #     dicta2 = {f: c for f, c in map(lambda x: x.strip().split(',')[1:], fp)}
    file_names_to_highlight = []
    with open("archive/data_images_2300.csv", "r") as fp:
        for line in csv.reader(fp):
            filename = line[0].strip()
            image_path = None
            if filename in train_filenames:
                image_path = (dataset / "train" / "JPEGImages" / filename)
            if filename in test_filenames:
                image_path = (dataset / "test" / "JPEGImages" / filename)
            
            if image_path:
                image_paths.append(image_path)
    with open("archive/data_images_500.csv", "r") as fp:
        for line in csv.reader(fp):
            filename = line[0].strip()
            file_names_to_highlight.append(os.path.basename(filename))
    # Example usage
    save_path = str(destination)
    # with open("caption_soumen.csv", "r") as fp:
    #     mark_file_list = [line[0].strip() for line in csv.reader(fp)]
    
    # generate_excel_with_thumbnails_pil(image_paths, save_path, mark_file_list, dicta, dicta2)
    generate_excel_with_thumbnails(image_paths, save_path, file_names_to_highlight)
    # generate_excel_form_csv_with_thumbnails(image_paths, save_path, "./dataset_500_to_2300_map_final.csv")
    print("Excel generated")


if __name__ == '__main__':
    main()
